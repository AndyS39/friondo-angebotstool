# Preislisten-Import (Phase 2, ab Phase 11 auf die EK-Preisliste umgestellt):
# - liest die TAIFUN-Preisliste v2 (11 Spalten, Erkennung über Header-Namen;
#   GUID-Anker, Kategoriezeilen, _x000D_-Bereinigung, "EP.", EK/Multi/Artikelnummer)
# - wendet die Textregeln aus der Logik-Excel v2 an (Blatt "Textregeln")
# - importiert die Zusatzartikel Z01–Z22 (Blatt "Zusatzartikel", inkl. EK Material);
#   bei "analog Pos. X" wird der Beschreibungstext übernommen und angepasst
# - Plausiprüfung: |EK × Multi − VK| > 1 € ergibt eine Hinweiszeile im Importbericht
# - Re-Import: Diff-Vorschau mit Warnliste (Positionsnummer <-> GUID-Abweichungen)

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from sqlalchemy.orm import Session

from app import config
from app.models import Artikel, QUELLE_PREISLISTE, QUELLE_ZUSATZ


# --- Textbereinigung ------------------------------------------------------

def text_bereinigen(wert) -> str:
    """Entfernt TAIFUN-Artefakte: _x000D_, weiche Trennzeichen, Leerzeichen-Padding."""
    if wert is None:
        return ""
    text = str(wert).replace("_x000D_", "").replace("\xad", "")
    zeilen = [re.sub(r" {2,}", " ", z).strip() for z in text.splitlines()]
    return "\n".join(z for z in zeilen if z).strip()


def kategorie_bereinigen(wert) -> str:
    """Kategorie-Überschrift: wie text_bereinigen, aber einzeilig."""
    return " ".join(text_bereinigen(wert).splitlines()).strip()


def euro_zu_cent(wert) -> int:
    return int((Decimal(str(wert)) * 100).quantize(Decimal("1")))


def datum_text(wert) -> str:
    """EK-Datum als deutscher Datumstext (Zelle kann Datum oder Text sein)."""
    if wert in (None, ""):
        return ""
    if isinstance(wert, (datetime, date)):
        return wert.strftime("%d.%m.%Y")
    return str(wert).strip()


def _header_indizes(kopfzeile, bericht_warnungen: list[str]) -> dict[str, int]:
    """Spalten über Header-Namen erkennen (Layout v2 ist nicht mehr fest).
    Die Einheiten-Spalte trägt in der TAIFUN-Datei keinen Namen und wird als
    erste unbenannte Spalte zwischen 'Menge' und 'Beschreibung' erkannt."""
    namen = [str(z).strip() if z is not None else "" for z in kopfzeile]
    indizes: dict[str, int] = {}
    for i, name in enumerate(namen):
        if name and name not in indizes:
            indizes[name] = i
    pflicht = ["GUID", "Position", "Menge", "Beschreibung", "E-Preis", "G-Preis"]
    for name in pflicht:
        if name not in indizes:
            raise ValueError(f"Preisliste: Pflichtspalte „{name}“ nicht im Header gefunden.")
    if "Einheit" not in indizes:
        for i in range(indizes["Menge"] + 1, indizes["Beschreibung"]):
            if not namen[i]:
                indizes["Einheit"] = i
                break
    if "Einheit" not in indizes:
        raise ValueError("Preisliste: Einheiten-Spalte nicht gefunden "
                         "(weder Header „Einheit“ noch unbenannte Spalte nach „Menge“).")
    for optional in ("Multi", "Artikelnummer", "Datum Einkaufspreis Material",
                     "Einkaufspreis Material"):
        if optional not in indizes:
            bericht_warnungen.append(
                f"Preisliste: Spalte „{optional}“ fehlt – Feld bleibt leer.")
    return indizes


# --- Textregeln aus der Logik-Excel --------------------------------------

@dataclass
class Textregeln:
    begriff_entfernen: dict[str, str] = field(default_factory=dict)   # pos_nr -> Begriff
    ueberschrift_ersetzen: dict[str, str] = field(default_factory=dict)
    ueberschrift_entfaellt: set[str] = field(default_factory=set)
    nicht_anwendbar: list[str] = field(default_factory=list)          # Warnungen


def lade_textregeln(wb_logik) -> Textregeln:
    regeln = Textregeln()
    for betrifft, regel, _bem in wb_logik["Textregeln"].iter_rows(min_row=2, values_only=True):
        betrifft = (str(betrifft or "")).strip()
        regel = (str(regel or "")).strip()
        if not betrifft or not regel:
            continue
        if betrifft == "Import allgemein":
            continue  # _x000D_-Bereinigung und Kategoriezeilen sind fest eingebaut
        m_pos = re.fullmatch(r"Position\s+(\w+)", betrifft)
        m_kopf = re.fullmatch(r"Überschrift\s+'(.+)'", betrifft)
        m_entf = re.search(r"Alle\s+'(.+)'-Nennungen.*entfernen", regel)
        m_ers = re.search(r"Ersetzen durch\s+'(.+)'", regel)
        if m_pos and m_entf:
            regeln.begriff_entfernen[m_pos.group(1)] = m_entf.group(1)
        elif m_kopf and m_ers:
            regeln.ueberschrift_ersetzen[m_kopf.group(1)] = m_ers.group(1)
        elif m_kopf and regel.lower().startswith("entfällt"):
            regeln.ueberschrift_entfaellt.add(m_kopf.group(1))
        else:
            regeln.nicht_anwendbar.append(
                f"Textregel nicht automatisch anwendbar: „{betrifft}“ – „{regel}“")
    return regeln


def begriff_entfernen(text: str, begriff: str) -> str:
    """Entfernt einen Begriff zeilenweise (Zeilenstruktur bleibt erhalten)."""
    zeilen = []
    for zeile in text.splitlines():
        zeile = re.sub(rf"[ ]*{re.escape(begriff)}[ ]*", " ", zeile)
        zeile = re.sub(r" {2,}", " ", zeile).strip()
        if zeile:
            zeilen.append(zeile)
    return "\n".join(zeilen)


# --- Preisliste + Zusatzartikel lesen ------------------------------------

@dataclass
class ImportErgebnis:
    artikel: list[dict] = field(default_factory=list)
    warnungen: list[str] = field(default_factory=list)


def lese_dateien() -> ImportErgebnis:
    """Liest Preisliste v2 und Zusatzartikel (Logik v2); liefert Artikel-Dicts + Warnungen."""
    ergebnis = ImportErgebnis()
    wb_logik = openpyxl.load_workbook(config.LOGIK_EXCEL_V2_PFAD, data_only=True)
    regeln = lade_textregeln(wb_logik)
    ergebnis.warnungen.extend(regeln.nicht_anwendbar)

    wb = openpyxl.load_workbook(config.PREISLISTE_PFAD, data_only=True)
    ws = wb[wb.sheetnames[0]]

    zeilen = ws.iter_rows(values_only=True)
    indizes = _header_indizes(next(zeilen), ergebnis.warnungen)

    def zelle(row, name):
        i = indizes.get(name)
        return row[i] if i is not None and i < len(row) else None

    kategorie = ""
    nach_pos: dict[str, dict] = {}
    for zeile_nr, row in enumerate(zeilen, start=2):
        guid = zelle(row, "GUID")
        beschreibung = zelle(row, "Beschreibung")
        pos = zelle(row, "Position")
        if guid is None and beschreibung is None:
            continue
        if pos is None:
            # Kategoriezeile (Gruppen-Überschrift) – Textregeln anwenden
            roh = kategorie_bereinigen(beschreibung)
            if roh in regeln.ueberschrift_entfaellt:
                continue  # Positionen laufen unter der vorherigen Überschrift weiter
            kategorie = regeln.ueberschrift_ersetzen.get(roh, roh)
            continue
        pos_nr = str(pos).strip()
        if guid is None:
            ergebnis.warnungen.append(f"Zeile {zeile_nr}: Position {pos_nr} ohne GUID – übersprungen.")
            continue
        text = text_bereinigen(beschreibung)
        if pos_nr in regeln.begriff_entfernen:
            text = begriff_entfernen(text, regeln.begriff_entfernen[pos_nr])
        g_preis = zelle(row, "G-Preis")
        e_preis = zelle(row, "E-Preis")
        ep_flag = isinstance(g_preis, str) and "EP" in g_preis
        if e_preis is None:
            ergebnis.warnungen.append(f"Position {pos_nr}: kein E-Preis – mit 0,00 € importiert.")
        ek = zelle(row, "Einkaufspreis Material")
        multi = zelle(row, "Multi")
        artikel = {
            "guid": str(guid).strip(),
            "pos_nr": pos_nr,
            "kategorie": kategorie,
            "bezeichnung": "",
            "beschreibung": text,
            "menge_standard": float(zelle(row, "Menge") or 1),
            "einheit": str(zelle(row, "Einheit") or "").strip(),
            "e_preis_cent": euro_zu_cent(e_preis or 0),
            "ep_flag": ep_flag,
            "quelle": QUELLE_PREISLISTE,
            "artikelnummer": str(zelle(row, "Artikelnummer") or "").strip(),
            "multi": float(multi) if multi not in (None, "") else None,
            "ek_cent": euro_zu_cent(ek) if ek not in (None, "") else None,
            "ek_datum": datum_text(zelle(row, "Datum Einkaufspreis Material")),
        }
        _plausibilitaet_pruefen(artikel, ergebnis)
        ergebnis.artikel.append(artikel)
        nach_pos[pos_nr] = artikel

    _zusatzartikel_lesen(wb_logik, nach_pos, ergebnis)
    return ergebnis


def _plausibilitaet_pruefen(artikel: dict, ergebnis: ImportErgebnis) -> None:
    """Hinweis, wenn EK × Multi um mehr als 1 € vom VK abweicht."""
    if artikel["ek_cent"] is None or not artikel["multi"]:
        return
    erwartet = artikel["ek_cent"] * artikel["multi"]
    abweichung_cent = abs(erwartet - artikel["e_preis_cent"])
    if abweichung_cent > 100:
        ergebnis.warnungen.append(
            f"Plausibilität Pos. {artikel['pos_nr']}: EK × Multi = "
            f"{erwartet / 100:,.2f} €, VK = {artikel['e_preis_cent'] / 100:,.2f} € "
            f"(Abweichung {abweichung_cent / 100:,.2f} €).".replace(",", "X")
            .replace(".", ",").replace("X", "."))


def _zusatzartikel_lesen(wb_logik, nach_pos: dict[str, dict], ergebnis: ImportErgebnis) -> None:
    """Zusatzartikel aus der Logik-Excel v2; Spalten über Header-Namen erkennen
    (u. a. 'VK netto (€)' und 'EK Material (€) – bitte ergänzen')."""
    zeilen = wb_logik["Zusatzartikel"].iter_rows(values_only=True)
    kopf = [str(z).strip() if z is not None else "" for z in next(zeilen)]

    def spalte(*muster):
        for i, name in enumerate(kopf):
            if any(m.lower() in name.lower() for m in muster):
                return i
        return None

    i_nr = spalte("Nr")
    i_bez = spalte("Bezeichnung")
    i_einheit = spalte("Einheit")
    i_vk = spalte("VK")
    i_ek = spalte("EK Material")
    i_text = spalte("Beschreibung", "Textquelle")
    if i_ek is None:
        ergebnis.warnungen.append(
            "Zusatzartikel: EK-Spalte nicht gefunden – Einkaufspreise bleiben leer.")

    for row in zeilen:
        nr = row[i_nr] if i_nr is not None else None
        if not nr:
            continue
        nr = str(nr).strip()
        bez = text_bereinigen(row[i_bez]) if i_bez is not None else ""
        quelle_text = (str(row[i_text] or "")).strip() if i_text is not None else ""
        ek = row[i_ek] if i_ek is not None else None
        beschreibung = ""
        m_analog = re.search(r"analog Pos\.?\s*(\d+)", quelle_text)
        if m_analog:
            basis = nach_pos.get(m_analog.group(1))
            if basis is None:
                ergebnis.warnungen.append(
                    f"{nr}: Textquelle Pos. {m_analog.group(1)} nicht in der Preisliste gefunden.")
            else:
                beschreibung = _text_anpassen(basis["beschreibung"], bez)
        elif quelle_text:
            beschreibung = quelle_text  # eigenständiger Beschreibungstext (z. B. Z22)
        ergebnis.artikel.append({
            "guid": None,
            "pos_nr": nr,
            "kategorie": "Zusatzartikel",
            "bezeichnung": bez,
            "beschreibung": beschreibung,
            "menge_standard": 1.0,
            "einheit": str((row[i_einheit] if i_einheit is not None else "") or "").strip(),
            "e_preis_cent": euro_zu_cent((row[i_vk] if i_vk is not None else 0) or 0),
            "ep_flag": False,
            "quelle": QUELLE_ZUSATZ,
            "artikelnummer": "",
            "multi": None,
            "ek_cent": euro_zu_cent(ek) if ek not in (None, "") else None,
            "ek_datum": "",
        })


def _text_anpassen(basis_text: str, ziel_bezeichnung: str) -> str:
    """Passt den übernommenen Beschreibungstext an Material/Größe des Z-Artikels an.

    - Tankgröße/-material: "bis 5.000 L Tank" -> "bis 3.000 L Kunststofftank"
    - Heizkörpergröße:     "Größe M"          -> "Größe S"
    """
    text = basis_text
    m_tank = re.search(r"bis\s+([\d.]+)\s*L", ziel_bezeichnung)
    if m_tank:
        material = "Tank"
        if "kunststoff" in ziel_bezeichnung.lower():
            material = "Kunststofftank"
        elif "stahl" in ziel_bezeichnung.lower():
            material = "Stahltank"
        text = re.sub(r"bis\s+[\d.]+\s*L\s*\w*[Tt]ank\w*",
                      f"bis {m_tank.group(1)} L {material}", text)
    m_groesse = re.search(r"Größe\s+(\w+)\s*$", ziel_bezeichnung)
    if m_groesse:
        text = re.sub(r"Größe\s+\w+", f"Größe {m_groesse.group(1)}", text)
    return text


# --- Diff-Vorschau und Übernahme ------------------------------------------

FELDER_VERGLEICH = ["pos_nr", "kategorie", "bezeichnung", "beschreibung",
                    "menge_standard", "einheit", "e_preis_cent", "ep_flag",
                    "artikelnummer", "multi", "ek_cent", "ek_datum"]

FELD_NAMEN = {
    "pos_nr": "Positionsnummer", "kategorie": "Kategorie", "bezeichnung": "Bezeichnung",
    "beschreibung": "Beschreibung", "menge_standard": "Standardmenge",
    "einheit": "Einheit", "e_preis_cent": "E-Preis", "ep_flag": "EP-Kennzeichen",
    "artikelnummer": "Artikelnummer", "multi": "Multi", "ek_cent": "EK Material",
    "ek_datum": "EK-Datum",
}


@dataclass
class Diff:
    neu: list[dict] = field(default_factory=list)
    geaendert: list[tuple] = field(default_factory=list)      # (Artikel, dict, [Feldnamen])
    unveraendert: int = 0
    entfallen: list = field(default_factory=list)             # Artikel, nicht mehr in Datei
    reaktiviert: list = field(default_factory=list)
    warnungen: list[str] = field(default_factory=list)


def _bestand_finden(daten: dict, nach_guid: dict, zusatz_nach_pos: dict):
    """Anker: GUID für Preislisten-Artikel, Pos-Nr. für Z-Artikel."""
    if daten["guid"]:
        return nach_guid.get(daten["guid"])
    return zusatz_nach_pos.get(daten["pos_nr"])


def berechne_diff(session: Session, ergebnis: ImportErgebnis) -> Diff:
    diff = Diff(warnungen=list(ergebnis.warnungen))
    bestand = session.query(Artikel).filter(Artikel.quelle != "manuell").all()
    nach_guid = {a.guid: a for a in bestand if a.guid}
    zusatz_nach_pos = {a.pos_nr: a for a in bestand if a.quelle == QUELLE_ZUSATZ}
    bestand_nach_pos = {a.pos_nr: a for a in bestand if a.pos_nr}

    gefunden_ids = set()
    for daten in ergebnis.artikel:
        vorhanden = _bestand_finden(daten, nach_guid, zusatz_nach_pos)

        # Warnliste: hinter einer Positionsnummer steckt jetzt eine andere GUID
        anderer = bestand_nach_pos.get(daten["pos_nr"])
        if (daten["guid"] and anderer is not None and anderer.guid
                and anderer.guid != daten["guid"]):
            diff.warnungen.append(
                f"Position {daten['pos_nr']}: andere GUID als bisher "
                f"(bisher „{anderer.titel[:60]}“, neu „{daten['beschreibung'].splitlines()[0][:60]}“). "
                "Bitte prüfen – Zuordnung erfolgt über die GUID.")

        if vorhanden is None:
            diff.neu.append(daten)
            continue
        gefunden_ids.add(vorhanden.id)
        felder = [f for f in FELDER_VERGLEICH if getattr(vorhanden, f) != daten[f]]
        if "pos_nr" in felder:
            diff.warnungen.append(
                f"Artikel „{vorhanden.titel[:60]}“: Positionsnummer wechselt "
                f"von {vorhanden.pos_nr} zu {daten['pos_nr']}.")
        if felder:
            diff.geaendert.append((vorhanden, daten, [FELD_NAMEN[f] for f in felder]))
        else:
            diff.unveraendert += 1
        if not vorhanden.aktiv:
            diff.reaktiviert.append(vorhanden)

    diff.entfallen = [a for a in bestand if a.id not in gefunden_ids and a.aktiv]
    for a in diff.entfallen:
        diff.warnungen.append(
            f"Pos. {a.pos_nr} „{a.titel[:60]}“ ist nicht mehr in der Datei – wird deaktiviert.")
    return diff


def import_ausfuehren(session: Session) -> tuple[Diff, str]:
    """Wendet den Import an (Diff wird direkt aus den Dateien neu berechnet)."""
    ergebnis = lese_dateien()
    diff = berechne_diff(session, ergebnis)

    nach_guid = {a.guid: a for a in session.query(Artikel).filter(Artikel.guid.isnot(None))}
    zusatz_nach_pos = {a.pos_nr: a for a in
                       session.query(Artikel).filter(Artikel.quelle == QUELLE_ZUSATZ)}
    for daten in ergebnis.artikel:
        vorhanden = _bestand_finden(daten, nach_guid, zusatz_nach_pos)
        if vorhanden is None:
            session.add(Artikel(**daten, aktiv=True))
        else:
            for feld in FELDER_VERGLEICH:
                setattr(vorhanden, feld, daten[feld])
            vorhanden.aktiv = True
    for artikel in diff.entfallen:
        artikel.aktiv = False
    session.commit()

    meldung = (f"{len(diff.neu)} neu, {len(diff.geaendert)} geändert, "
               f"{diff.unveraendert} unverändert, {len(diff.entfallen)} deaktiviert")
    return diff, meldung
