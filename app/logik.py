# Logik-Import (Phase 3): Parser für konfigurator_logik.xlsx
# (Blätter Fragen, Aktionen, Paketmatrix, Angebotsaufbau, KfW) plus Validierung.
# Die geparste Logik wird im Prozess gecacht; "Konfiguration neu einlesen"
# ersetzt den Cache. Die inhaltliche Auswertung (Fragenfluss, KfW-Rechnung)
# folgt in den Phasen 4–6 – hier geht es um Struktur, Referenzen, Bedingungen.

import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

from app import config
from app.models import Artikel

FRAGE_TYPEN = {
    "Auswahl",
    "Zahleneingabe",
    "Betragseingabe",
    "Mengenmaske",
    "Wiederholfeld",
    "Freitext",
    "Freitext groß",
    "Datum",            # v8: Datumseingabe (z. B. Wiedervorlage der Einschätzung)
}

# Alt-/Varianten-Schreibweisen werden beim Einlesen normalisiert
TYP_ALIASE = {
    "Mengenmaske (4 Zahlenfelder)": "Mengenmaske",
    "Mengenmaske (2 Zahlenfelder)": "Mengenmaske",
    "Wiederholfeld je Verteiler": "Wiederholfeld",
}

ZAHLEN_TYPEN = {"Zahleneingabe", "Betragseingabe", "Mengenmaske", "Wiederholfeld"}

FREITEXT_TYPEN = {"Freitext", "Freitext groß"}

# Aktionszeilen, die keine einzelne Frage betreffen
SPEZIAL_AKTIONEN = {"Gruppen-Trigger", "Grundpaket", "Ampel-Auswertung",
                    "O01–O08 / K01–K04", "O01-O08 / K01-K04"}


# --- Datenstrukturen ------------------------------------------------------

@dataclass
class Bedingung:
    roh: str
    # immer | antwort | ausgefuellt | selbstnutzung | friondo_ja | klauseln |
    # wiederholgruppe (v8: „je Raum (KO05)“ – Frage wiederholt sich je Zähler)
    art: str
    frage_id: Optional[str] = None
    werte: list[str] = field(default_factory=list)
    # v3: ODER-verknüpfte Klauseln, jede Klausel = UND-Liste von (frage_id, werte),
    # z. B. "nur wenn A04 = KG oder EG, oder (A04 = DG und D01 = Nein)"
    klauseln: list[list[tuple[str, list[str]]]] = field(default_factory=list)


@dataclass
class Frage:
    id: str
    reihenfolge: int
    text: str
    typ: str
    antworten: list[str]
    bedingung: Optional[Bedingung]
    hinweis: str
    seite: str = ""          # Kategorie-Seite der mobilen Erfassung (v2)


@dataclass
class ArtikelRef:
    ref: str                      # "045" oder "Z01"
    menge: str = "1"              # roh: "1", "2", "eingegebene Meter", "Anzahl Verteiler", ...
    ep: bool = False
    kein_ep: bool = False         # v8: "(kein EP)" – überschreibt das EP-Flag des Artikelstamms


@dataclass
class Aktion:
    frage: str                    # "A01" oder Spezialschlüssel (Gruppen-Trigger, Grundpaket, ...)
    antwort: str                  # roh, z. B. "Gas", "Kunststoff, bis 3.000 L", "50 l / 100 l / 200 l"
    aktion_roh: str
    typ: str                      # ampel | normal
    ampel_grund: str              # Klartext-Grund bei AMPEL-Antworten (v2: kein Abbruch)
    artikel: list[ArtikelRef]
    bemerkung: str


@dataclass
class PaketZeile:
    leistungsklasse: str
    verbrauch_roh: str
    verbrauch_von: Optional[int]
    verbrauch_bis: Optional[int]
    ww_bis_200: list[ArtikelRef]
    ohne_ww: list[ArtikelRef]
    ww_300: list[ArtikelRef]
    # v8: Heizlast-Spalte (kW, Dezimalzahlen) – hat bei Bekanntsein Vorrang
    heizlast_roh: str = ""
    heizlast_von: Optional[float] = None
    heizlast_bis: Optional[float] = None


@dataclass
class AngebotsBlock:
    nr: int
    ueberschrift: str
    inhalt_roh: str
    wann: Optional[Bedingung]
    refs: list[ArtikelRef]


@dataclass
class Anhang:
    datei: str                              # Dateiname im Ordner anlagen/
    regel_roh: str
    art: str                                # immer | frage | position | unbekannt
    frage_id: str = ""
    antwort: str = ""
    positionen: list[str] = field(default_factory=list)
    bemerkung: str = ""


@dataclass
class Logik:
    fragen: dict[str, Frage]
    aktionen: list[Aktion]
    pakete: list[PaketZeile]
    bloecke: list[AngebotsBlock]
    kfw: dict[str, tuple[str, str]]         # Parameter -> (Wert, Bemerkung)
    geladen_am: datetime
    anhaenge: list[Anhang] = field(default_factory=list)
    # v8: reine Erfassungsbögen je Sparte (Blätter "Fragen PV" / "Fragen KL"),
    # ohne Artikel-Aktionen – laufen über die TAIFUN-Schiene
    sparten_fragen: dict[str, dict[str, Frage]] = field(default_factory=dict)

    @property
    def seiten(self) -> list[str]:
        """Seiten der mobilen Erfassung in Blatt-Reihenfolge."""
        ergebnis: list[str] = []
        for frage in sorted(self.fragen.values(), key=lambda f: f.reihenfolge):
            if frage.seite and frage.seite not in ergebnis:
                ergebnis.append(frage.seite)
        return ergebnis


@dataclass
class Pruefbericht:
    fehler: list[str] = field(default_factory=list)
    warnungen: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.fehler


# --- Referenzen und Bedingungen parsen ------------------------------------

def refs_extrahieren(text: str) -> list[ArtikelRef]:
    """Extrahiert Artikel-Referenzen: 'Pos. 045', Listen ('Pos. 003, 005, 007 (EP)'),
    Slash-Listen ('Pos. 149 / 150 / 153'), Z-Artikel inkl. Bereichen ('Z01–Z14')."""
    if not text:
        return []
    refs: list[ArtikelRef] = []

    menge_muster = r"\s*×\s*(\([^)]*\)|[\wäöüÄÖÜß. ]+)"   # auch "(Eingabe − 3)"
    gefunden: list[tuple[int, ArtikelRef]] = []

    for m in re.finditer(r"Pos\.\s*((?:\d{1,3}(?:\s*\(EP\))?\s*(?:[,/]\s*)?)+)", text):
        nummern = re.findall(r"(\d{1,3})(\s*\(EP\))?", m.group(1))
        rest = text[m.end():]
        m_menge = re.match(menge_muster, rest)
        menge = m_menge.group(1).strip() if m_menge else "1"
        menge = re.sub(r"\s*als EP.*$", "", menge).strip() or "1"
        ep_nach = bool(re.match(r"[^+·]*als EP", rest))
        kein_ep = bool(re.match(r"[^+·]*\(kein EP\)", rest))   # v8: EP-Flag unterdrücken
        for i, (nummer, ep) in enumerate(nummern):
            gefunden.append((m.start() + i,
                             ArtikelRef(nummer.zfill(3), menge, bool(ep) or ep_nach,
                                        kein_ep)))

    for m in re.finditer(r"\bZ(\d{2})\b(?:\s*[–-]\s*Z(\d{2}))?", text):
        von, bis = int(m.group(1)), int(m.group(2) or m.group(1))
        rest = text[m.end():]
        m_menge = re.match(menge_muster, rest)
        menge = m_menge.group(1).strip() if (m_menge and von == bis) else "1"
        for n in range(von, bis + 1):
            gefunden.append((m.start() + (n - von),
                             ArtikelRef(f"Z{n:02d}", menge, False)))

    # Reihenfolge wie im Text – wichtig für die paarweise Zuordnung zu Slash-Listen
    gefunden.sort(key=lambda t: t[0])
    refs.extend(ref for _, ref in gefunden)
    return refs


def bedingung_parsen(roh) -> Optional[Bedingung]:
    """Parst 'Anzeigen wenn' / 'Wann'; None bei nicht parsebarer Bedingung."""
    text = str(roh or "").strip()
    if text in ("", "immer", "–", "-"):
        return Bedingung(text, "immer")
    if re.match(r"nur bei Selbstnutzung", text):
        return Bedingung(text, "selbstnutzung")
    if re.search(r"Friondo-Ja", text):
        return Bedingung(text, "friondo_ja")
    # v8: Wiederholgruppe – „je Raum (KO05)“: Fragen wiederholen sich so oft,
    # wie die referenzierte Zählfrage angibt
    m = re.match(r"je\s+\S+\s*\(([A-Z]{1,2}\d{2})\)$", text)
    if m:
        return Bedingung(text, "wiederholgruppe", m.group(1))
    m = re.match(r"nur wenn ([A-Z]{1,2}\d{2})\s+ausgefüllt$", text)
    if m:
        return Bedingung(text, "ausgefuellt", m.group(1))
    m = re.match(r"nur wenn (.+)$", text)
    if m:
        # ", oder " trennt ODER-Klauseln; innerhalb einer Klausel trennt " und ";
        # " oder " ohne Komma trennt Werte derselben Frage ("KG oder EG")
        klauseln: list[list[tuple[str, list[str]]]] = []
        for klausel_text in re.split(r",\s*oder\s+", m.group(1)):
            klausel_text = klausel_text.strip()
            if klausel_text.startswith("(") and klausel_text.endswith(")"):
                klausel_text = klausel_text[1:-1].strip()
            terme: list[tuple[str, list[str]]] = []
            for teil in re.split(r"\s+und\s+", klausel_text):
                tm = re.match(r"([A-Z]{1,2}\d{2})\s*=\s*(.+)$", teil.strip())
                if tm is None:
                    return None
                terme.append((tm.group(1),
                              [w.strip() for w in tm.group(2).split(" oder ")]))
            klauseln.append(terme)
        if len(klauseln) == 1 and len(klauseln[0]) == 1:
            frage_id, werte = klauseln[0][0]
            return Bedingung(text, "antwort", frage_id, werte)
        return Bedingung(text, "klauseln", klauseln=klauseln)
    return None


def _alias_aufloesen(wert: str, optionen: list[str]) -> Optional[str]:
    """Antwort-Alias auf eine Option abbilden:
    - 'beides' -> Option mit ' und '
    - eindeutiger Präfix, z. B. 'Mehrfamilienhaus' -> 'Mehrfamilienhaus (2+ WE)'"""
    if wert in optionen:
        return wert
    if wert.lower() == "beides":
        for option in optionen:
            if " und " in option:
                return option
    praefix_treffer = [o for o in optionen if o.startswith(wert)]
    if len(praefix_treffer) == 1:
        return praefix_treffer[0]
    return None


# --- Excel einlesen -------------------------------------------------------

def _zelle(wert) -> str:
    return str(wert).strip() if wert is not None else ""


def logik_einlesen() -> tuple[Logik, Pruefbericht]:
    bericht = Pruefbericht()
    wb = openpyxl.load_workbook(config.LOGIK_EXCEL_PFAD, data_only=True)

    for blatt in ("Fragen", "Aktionen", "Paketmatrix", "Angebotsaufbau", "KfW"):
        if blatt not in wb.sheetnames:
            bericht.fehler.append(f"Blatt „{blatt}“ fehlt in der Logik-Excel.")
    if bericht.fehler:
        return Logik({}, [], [], [], {}, datetime.now()), bericht

    fragen = _fragen_einlesen(wb, bericht)
    aktionen = _aktionen_einlesen(wb, bericht)
    pakete = _paketmatrix_einlesen(wb, bericht)
    bloecke = _angebotsaufbau_einlesen(wb, bericht)
    kfw = _kfw_einlesen(wb, bericht)
    anhaenge = _anhaenge_einlesen(wb, bericht)

    # v8: reine Erfassungsbögen PV/KL (eigene Blätter, ohne Artikel-Aktionen)
    sparten_fragen: dict[str, dict[str, Frage]] = {}
    for sparte, blatt in (("PV", "Fragen PV"), ("KL", "Fragen KL")):
        if blatt not in wb.sheetnames:
            bericht.warnungen.append(
                f"Blatt „{blatt}“ fehlt – die {sparte}-Erfassung läuft nur als Freitext.")
            continue
        sparten_fragen[sparte] = _fragen_einlesen(wb, bericht, blatt)

    logik = Logik(fragen, aktionen, pakete, bloecke, kfw, datetime.now(), anhaenge,
                  sparten_fragen)
    _querbezuege_pruefen(logik, bericht)
    for sparte, sfragen in sparten_fragen.items():
        _bedingungen_pruefen(sfragen, bericht, f"Fragen {sparte}")
    return logik, bericht


def _anhaenge_einlesen(wb, bericht: Pruefbericht) -> list[Anhang]:
    """Blatt "Anhänge": Datei · Regel · Bemerkung.
    Regeln: 'immer' | 'wenn <Frage> = <Antwort>' | 'wenn Pos. <Nr> im Angebot'."""
    if "Anhänge" not in wb.sheetnames:
        bericht.warnungen.append("Blatt „Anhänge“ fehlt – es werden keine Anhänge geregelt.")
        return []
    anhaenge = []
    for datei, regel, bemerkung in wb["Anhänge"].iter_rows(min_row=2, values_only=True):
        datei = _zelle(datei)
        regel = _zelle(regel)
        if not datei or datei.startswith("("):
            continue  # Platzhalterzeile
        eintrag = Anhang(datei, regel, "unbekannt", bemerkung=_zelle(bemerkung))
        if regel == "immer":
            eintrag.art = "immer"
        elif (m := re.match(r"wenn\s+([A-Z]\d{2})\s*=\s*(.+)$", regel)):
            eintrag.art = "frage"
            eintrag.frage_id = m.group(1)
            eintrag.antwort = m.group(2).strip()
        elif (m := re.match(r"wenn\s+(?:WP-Paket\s+)?Pos\.\s*([\d–\-\s]+)\s+im Angebot", regel)):
            eintrag.art = "position"
            teil = m.group(1).strip()
            m_bereich = re.match(r"(\d{1,3})\s*[–-]\s*(\d{1,3})$", teil)
            if m_bereich:
                eintrag.positionen = [f"{n:03d}" for n in
                                      range(int(m_bereich.group(1)), int(m_bereich.group(2)) + 1)]
            else:
                eintrag.positionen = [n.zfill(3) for n in re.findall(r"\d{1,3}", teil)]
        else:
            bericht.warnungen.append(f"Anhänge: Regel „{regel}“ für {datei} nicht lesbar.")
        anhaenge.append(eintrag)
    return anhaenge


def _fragen_einlesen(wb, bericht: Pruefbericht, blatt: str = "Fragen") -> dict[str, Frage]:
    """v2-Layout: Seite · ID · Fragetext · Typ · Antworten · Anzeigen wenn · Hinweis.
    Die Reihenfolge ergibt sich aus der Zeilenfolge im Blatt. Seit v8 auch für
    die Sparten-Blätter „Fragen PV“ / „Fragen KL“."""
    fragen: dict[str, Frage] = {}
    for zeile, row in enumerate(wb[blatt].iter_rows(min_row=2, values_only=True), 2):
        seite, fid, text, typ, antworten, anzeigen, hinweis = (_zelle(v) for v in row[:7])
        if not fid:
            continue
        if fid in fragen:
            bericht.fehler.append(f"{blatt} Zeile {zeile}: ID {fid} doppelt vergeben.")
            continue
        typ = TYP_ALIASE.get(typ, typ)
        if typ not in FRAGE_TYPEN:
            bericht.fehler.append(f"{blatt} {fid}: unbekannter Typ „{typ}“.")
        if not seite:
            bericht.fehler.append(f"{blatt} {fid}: Spalte „Seite“ ist leer.")
        optionen = [a.strip() for a in antworten.split("|") if a.strip()] if antworten else []
        if typ == "Auswahl" and not optionen:
            bericht.fehler.append(f"{blatt} {fid}: Auswahl ohne Antwortmöglichkeiten.")
        bedingung = bedingung_parsen(anzeigen)
        if bedingung is None:
            bericht.fehler.append(
                f"{blatt} {fid}: Bedingung „{anzeigen}“ nicht parsebar.")
        fragen[fid] = Frage(fid, len(fragen) + 1, text, typ, optionen, bedingung,
                            hinweis, seite)
    return fragen


def _aktionen_einlesen(wb, bericht: Pruefbericht) -> list[Aktion]:
    aktionen = []
    for row in wb["Aktionen"].iter_rows(min_row=2, values_only=True):
        frage, antwort, aktion_roh, bemerkung = (_zelle(v) for v in row[:4])
        if not frage:
            continue
        if aktion_roh.startswith("AMPEL"):
            # "AMPEL: individuell – Grund: <Text>" → Flag statt Abbruch
            typ = "ampel"
            m = re.search(r"Grund:\s*(.+)$", aktion_roh)
            grund = m.group(1).strip() if m else aktion_roh
        else:
            typ = "normal"
            grund = ""
        aktionen.append(Aktion(frage, antwort, aktion_roh, typ, grund,
                               refs_extrahieren(aktion_roh), bemerkung))
    return aktionen


def _paketmatrix_einlesen(wb, bericht: Pruefbericht) -> list[PaketZeile]:
    pakete = []
    for row in wb["Paketmatrix"].iter_rows(min_row=2, values_only=True):
        klasse, verbrauch, ww200, ohne_ww, ww300 = (_zelle(v) for v in row[:5])
        heizlast = _zelle(row[5]) if len(row) > 5 else ""
        if not klasse:
            continue
        von = bis = None
        m = re.match(r"([\d.]+)\s*[–-]\s*([\d.]+)\s*kWh", verbrauch)
        if m:
            von = int(m.group(1).replace(".", ""))
            bis = int(m.group(2).replace(".", ""))
        else:
            bericht.fehler.append(
                f"Paketmatrix {klasse}: Verbrauchsbereich „{verbrauch}“ nicht parsebar.")
        # v8: Heizlast-Spalte „bis 5,9 kW“ / „6,0 – 7,9 kW“ (Dezimal mit Komma)
        h_von = h_bis = None
        if heizlast:
            def _dez(t):
                return float(t.replace(".", "").replace(",", "."))
            if (m := re.match(r"([\d.,]+)\s*[–-]\s*([\d.,]+)", heizlast)):
                h_von, h_bis = _dez(m.group(1)), _dez(m.group(2))
            elif (m := re.match(r"bis\s*([\d.,]+)", heizlast)):
                h_von, h_bis = 0.0, _dez(m.group(1))
            else:
                bericht.fehler.append(
                    f"Paketmatrix {klasse}: Heizlast „{heizlast}“ nicht parsebar.")
        zeile = PaketZeile(klasse, verbrauch, von, bis,
                           refs_extrahieren(ww200), refs_extrahieren(ohne_ww),
                           refs_extrahieren(ww300), heizlast, h_von, h_bis)
        for spalte, refs in (("WW bis 200 l", zeile.ww_bis_200),
                             ("ohne Warmwasser", zeile.ohne_ww), ("WW 300 l", zeile.ww_300)):
            if not refs:
                bericht.fehler.append(
                    f"Paketmatrix {klasse}: Spalte „{spalte}“ ohne Artikel-Referenz.")
        pakete.append(zeile)
    return pakete


def _angebotsaufbau_einlesen(wb, bericht: Pruefbericht) -> list[AngebotsBlock]:
    bloecke = []
    for row in wb["Angebotsaufbau"].iter_rows(min_row=2, values_only=True):
        nr, ueberschrift, inhalt, wann = (_zelle(v) for v in row[:4])
        if not nr:
            continue
        if nr == "Nachtexte":
            continue  # Vollmacht-Bedingung (Nachtext D) – wird in pdf_export ausgewertet
        try:
            block_nr = int(float(nr))
        except ValueError:
            bericht.fehler.append(f"Angebotsaufbau: Blocknummer „{nr}“ keine Zahl.")
            continue
        bedingung = bedingung_parsen(wann)
        if bedingung is None:
            bericht.fehler.append(
                f"Angebotsaufbau Block {block_nr}: Bedingung „{wann}“ nicht parsebar.")
        refs = refs_extrahieren(inhalt)
        # v2-Schreibweise: nach "Pos. 005 · 006 · 007 (EP)" stehen weitere Nummern
        # ohne "Pos."-Präfix – nackte dreistellige Token ergänzen
        vorhanden = {r.ref for r in refs}
        for m in re.finditer(r"\b(\d{3})\b", inhalt):
            if m.group(1) not in vorhanden:
                refs.append(ArtikelRef(m.group(1), "1", False))
                vorhanden.add(m.group(1))
        bloecke.append(AngebotsBlock(block_nr, ueberschrift, inhalt, bedingung, refs))
    return bloecke


PFLICHT_KFW_PARAMETER = [
    "Gültigkeit der Konditionen", "Grundförderung", "Klimageschwindigkeits-Bonus",
    "Einkommensbonus", "Fördersatz-Deckel",
    "Höchstkosten EFH", "Höchstkosten MFH", "Höchstkosten Gewerbe",
    "ABLEITUNG Gebäudetyp", "ABLEITUNG Klima-Vorbelegung (K02)",
]


def _kfw_einlesen(wb, bericht: Pruefbericht) -> dict[str, tuple[str, str]]:
    kfw = {}
    for row in wb["KfW"].iter_rows(min_row=2, values_only=True):
        parameter, wert, bemerkung = (_zelle(v) for v in row[:3])
        if parameter:
            kfw[parameter] = (wert, bemerkung)
    for pflicht in PFLICHT_KFW_PARAMETER:
        if pflicht not in kfw:
            bericht.fehler.append(f"KfW: Parameter „{pflicht}“ fehlt.")
    gueltigkeit = kfw.get("Gültigkeit der Konditionen", ("", ""))[0]
    if gueltigkeit and not re.search(r"\d{2}\.\d{2}\.\d{4}\s*bis\s*\d{2}\.\d{2}\.\d{4}",
                                     gueltigkeit):
        bericht.warnungen.append(
            f"KfW: Gültigkeit „{gueltigkeit}“ nicht als Zeitraum (TT.MM.JJJJ bis TT.MM.JJJJ) lesbar.")
    return kfw


# --- Querbezüge validieren ------------------------------------------------

def antwort_teile(text: str) -> list[str]:
    """Zerlegt Slash-/oder-Listen aus Aktionszeilen. Slash nur mit Leerzeichen
    ringsum trennen, damit Optionswerte wie 'Luft/Wasser' erhalten bleiben."""
    return [t.strip() for t in re.split(r"\s+oder\s+|\s+/\s+", text) if t.strip()]


def _antwort_pruefen(frage: Frage, antwort: str, fragen: dict[str, Frage]) -> Optional[str]:
    """Prüft, ob eine Antwort/Bedingung aus dem Blatt Aktionen zur Frage passt.
    Liefert einen Fehlertext oder None."""
    if frage.typ == "Mengenmaske":
        # z. B. "Anzahl S / M / L / XL": erster Teil trägt das "Anzahl"-Präfix
        teile = antwort_teile(antwort)
        groessen = [re.sub(r"^Anzahl\s+", "", t) for t in teile]
        if groessen and all(g in frage.antworten for g in groessen):
            return None
        return f"„{antwort}“ passt nicht zur Mengenmaske ({' | '.join(frage.antworten)})."
    if frage.typ in ZAHLEN_TYPEN or frage.typ in FREITEXT_TYPEN:
        return None  # freie Beschreibungen/Bereiche erlaubt

    # Auswahl: "A", "A oder B", "A / B / C", "Bedingungsfrage-Wert, eigener Wert"
    teile = antwort_teile(antwort)
    if all(_alias_aufloesen(t, frage.antworten) for t in teile):
        return None
    if "," in antwort:
        vorne, hinten = (t.strip() for t in antwort.split(",", 1))
        eltern_optionen = []
        if frage.bedingung and frage.bedingung.frage_id in fragen:
            eltern_optionen = fragen[frage.bedingung.frage_id].antworten
        if (_alias_aufloesen(hinten, frage.antworten)
                and _alias_aufloesen(vorne, eltern_optionen)):
            return None
    return f"Antwort „{antwort}“ ist keine bekannte Option ({' | '.join(frage.antworten)})."


def _bedingungen_pruefen(fragen: dict[str, Frage], bericht: Pruefbericht,
                         kontext: str = "Fragen") -> None:
    """Bedingungen eines Fragen-Satzes: referenzierte Fragen + Werte müssen
    existieren (v8: auch für die Sparten-Blätter, inkl. Wiederholgruppen)."""
    for frage in fragen.values():
        b = frage.bedingung
        if b is None:
            continue
        terme: list[tuple[str, list[str]]] = []
        if b.art in ("antwort",):
            terme = [(b.frage_id, b.werte)]
        elif b.art in ("ausgefuellt", "wiederholgruppe"):
            terme = [(b.frage_id, [])]
        elif b.art == "klauseln":
            terme = [t for klausel in b.klauseln for t in klausel]
        for frage_id, werte in terme:
            if frage_id not in fragen:
                bericht.fehler.append(
                    f"{kontext} {frage.id}: Bedingung verweist auf unbekannte Frage {frage_id}.")
                continue
            ziel = fragen[frage_id]
            if ziel.typ != "Auswahl":
                continue
            for wert in werte:
                if _alias_aufloesen(wert, ziel.antworten) is None:
                    bericht.fehler.append(
                        f"{kontext} {frage.id}: Bedingungswert „{wert}“ ist keine Option von {ziel.id}.")


def _querbezuege_pruefen(logik: Logik, bericht: Pruefbericht) -> None:
    fragen = logik.fragen
    _bedingungen_pruefen(fragen, bericht)

    # Anhänge: referenzierte Fragen/Antworten müssen existieren
    for anhang in logik.anhaenge:
        if anhang.art != "frage":
            continue
        if anhang.frage_id not in fragen:
            bericht.fehler.append(
                f"Anhänge {anhang.datei}: unbekannte Frage {anhang.frage_id}.")
        elif (fragen[anhang.frage_id].typ == "Auswahl"
              and _alias_aufloesen(anhang.antwort, fragen[anhang.frage_id].antworten) is None):
            bericht.fehler.append(
                f"Anhänge {anhang.datei}: „{anhang.antwort}“ ist keine Option von {anhang.frage_id}.")

    # Aktionen: Frage bekannt, Antwort plausibel
    for aktion in logik.aktionen:
        if aktion.frage in SPEZIAL_AKTIONEN:
            continue
        if aktion.frage not in fragen:
            bericht.fehler.append(
                f"Aktionen: unbekannte Frage „{aktion.frage}“ (Antwort „{aktion.antwort}“).")
            continue
        problem = _antwort_pruefen(fragen[aktion.frage], aktion.antwort, fragen)
        if problem:
            bericht.fehler.append(f"Aktionen {aktion.frage}: {problem}")

    # Abgedeckte Antworten: jede Auswahl-Option sollte eine Aktionszeile haben.
    # Sammelzeilen wie "O01–O08 / K01–K04" decken ganze ID-Bereiche ab.
    sammelbereich: set[str] = set()
    for aktion in logik.aktionen:
        for praefix, von, bis in re.findall(r"([A-Z])(\d{2})[–-]\1(\d{2})", aktion.frage):
            sammelbereich.update(f"{praefix}{n:02d}"
                                 for n in range(int(von), int(bis) + 1))
    for frage in fragen.values():
        if frage.typ != "Auswahl" or frage.id in sammelbereich:
            continue
        # Kombinationslogik (z. B. A09 „Stahl, bis 5.000 L“ = Antwort der
        # Bedingungsfrage A08 + eigene Antwort): die Zeile deckt die eigene
        # Option nur für diese Eltern-Option ab – geprüft wird die ganze Matrix.
        eltern = (fragen.get(frage.bedingung.frage_id)
                  if frage.bedingung and frage.bedingung.frage_id else None)
        eltern_optionen = (frage.bedingung.werte or eltern.antworten) if eltern else []
        abgedeckt: set[str] = set()                       # Optionen ohne Eltern-Bezug
        kombi: dict[str, set[str]] = {}                   # Eltern-Option -> eigene Optionen
        for aktion in logik.aktionen:
            if aktion.frage != frage.id:
                continue
            kombiniert = False
            if "," in aktion.antwort and eltern is not None:
                vorne, hinten = (x.strip() for x in aktion.antwort.split(",", 1))
                e_opt = _alias_aufloesen(vorne, eltern.antworten)
                o_opt = _alias_aufloesen(hinten, frage.antworten)
                if e_opt and o_opt:
                    kombi.setdefault(e_opt, set()).add(o_opt)
                    kombiniert = True
            if kombiniert:
                continue
            for teil in [aktion.antwort] + antwort_teile(aktion.antwort):
                option = _alias_aufloesen(teil, frage.antworten)
                if option:
                    abgedeckt.add(option)
        if kombi:
            # je Eltern-Option, unter der die Frage gestellt wird, muss jede
            # eigene Option entweder kombiniert oder allgemein abgedeckt sein
            for e_opt in eltern_optionen:
                fehlend = [o for o in frage.antworten
                           if o not in abgedeckt and o not in kombi.get(e_opt, set())]
                if fehlend:
                    bericht.warnungen.append(
                        f"Aktionen {frage.id}: keine Aktionszeile für „{e_opt}, …“ bei "
                        f"Option(en) {', '.join(fehlend)}.")
            continue
        fehlend = [o for o in frage.antworten if o not in abgedeckt]
        if fehlend:
            bericht.warnungen.append(
                f"Aktionen {frage.id}: keine Aktionszeile für Option(en) {', '.join(fehlend)}.")


def artikel_referenzen(logik: Logik) -> dict[str, list[str]]:
    """Alle referenzierten Artikel mit Fundstellen (für die Validierung gegen die DB)."""
    fundstellen: dict[str, list[str]] = {}

    def merken(refs: list[ArtikelRef], quelle: str):
        for ref in refs:
            fundstellen.setdefault(ref.ref, []).append(quelle)

    for aktion in logik.aktionen:
        merken(aktion.artikel, f"Aktionen {aktion.frage} („{aktion.antwort}“)")
        merken(refs_extrahieren(aktion.bemerkung), f"Aktionen {aktion.frage} (Bemerkung)")
    for paket in logik.pakete:
        merken(paket.ww_bis_200, f"Paketmatrix {paket.leistungsklasse}")
        merken(paket.ohne_ww, f"Paketmatrix {paket.leistungsklasse}")
        merken(paket.ww_300, f"Paketmatrix {paket.leistungsklasse}")
    for block in logik.bloecke:
        merken(block.refs, f"Angebotsaufbau Block {block.nr}")
    return fundstellen


def artikel_pruefen(logik: Logik, session: Session, bericht: Pruefbericht) -> None:
    """Prüft, ob alle referenzierten Positionen/Z-Artikel als aktive Artikel existieren."""
    vorhanden = {pos for (pos,) in session.query(Artikel.pos_nr)
                 .filter(Artikel.aktiv.is_(True)) if pos}
    if not vorhanden:
        bericht.warnungen.append(
            "Artikelstamm ist leer – bitte zuerst die Preisliste importieren "
            "(Artikel → Preisliste importieren).")
        return
    for ref, quellen in sorted(artikel_referenzen(logik).items()):
        if ref not in vorhanden:
            bericht.fehler.append(
                f"Artikel {('Pos. ' + ref) if not ref.startswith('Z') else ref} "
                f"fehlt im Artikelstamm – referenziert in: {'; '.join(sorted(set(quellen)))}.")


# --- Cache / öffentliche API ---------------------------------------------

_cache: dict = {"logik": None, "bericht": None}


def neu_einlesen(session: Session) -> tuple[Logik, Pruefbericht]:
    logik, bericht = logik_einlesen()
    artikel_pruefen(logik, session, bericht)
    _cache["logik"], _cache["bericht"] = logik, bericht
    return logik, bericht


def hole_logik(session: Session) -> tuple[Logik, Pruefbericht]:
    if _cache["logik"] is None:
        return neu_einlesen(session)
    return _cache["logik"], _cache["bericht"]
